"""
Report Fixer Module - Transform Excel Reports to Match Template Layout

This module transforms Excel job reports (e.g., retech_jobRpt.xls) to match
a template layout. It tracks schedule changes, adds notes for date modifications,
and exports formatted Excel files with highlighting.
"""

import sys
import json
from pathlib import Path
from datetime import datetime
from PyQt6.QtWidgets import (
    QWidget, QTableWidgetItem, QFileDialog, QHeaderView, QAbstractItemView
)
from PyQt6.QtCore import Qt
from PyQt6 import uic

from core.base_module import BaseModule
from shared.utils import get_config_dir, open_folder

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportingModule(BaseModule):
    """Module for fixing and transforming Excel job reports"""

    def __init__(self):
        super().__init__()
        self._widget = None

        # Widget references
        self.template_path_edit = None
        self.source_path_edit = None
        self.source_info_label = None
        self.customer_combo = None
        self.preview_table = None
        self.mapping_status_label = None
        self.fix_export_btn = None
        self.open_output_btn = None
        self.status_text = None

        # State
        self.source_df = None
        self.template_columns = None
        self.last_output_path = None

    def get_name(self) -> str:
        return "Report Fixer"

    def get_order(self) -> int:
        return 80

    def is_experimental(self) -> bool:
        return False

    def initialize(self, app_context):
        super().initialize(app_context)

    def get_widget(self) -> QWidget:
        if self._widget is None:
            self._widget = self._create_widget()
        return self._widget

    def _create_widget(self) -> QWidget:
        """Create the report fixer tab widget"""
        widget = QWidget()

        # Check dependencies
        if not PANDAS_AVAILABLE or not OPENPYXL_AVAILABLE:
            from PyQt6.QtWidgets import QVBoxLayout, QLabel
            layout = QVBoxLayout(widget)
            missing = []
            if not PANDAS_AVAILABLE:
                missing.append("pandas")
            if not OPENPYXL_AVAILABLE:
                missing.append("openpyxl")
            label = QLabel(f"Missing required packages: {', '.join(missing)}\n\n"
                          f"Install with: pip install {' '.join(missing)}")
            label.setStyleSheet("color: red; padding: 20px;")
            layout.addWidget(label)
            return widget

        # Load UI file
        ui_file = self._get_ui_path('reporting/ui/reporting_tab.ui')
        uic.loadUi(ui_file, widget)

        # Store widget references
        self.template_path_edit = widget.template_path_edit
        self.source_path_edit = widget.source_path_edit
        self.source_info_label = widget.source_info_label
        self.customer_combo = widget.customer_combo
        self.preview_table = widget.preview_table
        self.mapping_status_label = widget.mapping_status_label
        self.fix_export_btn = widget.fix_export_btn
        self.open_output_btn = widget.open_output_btn
        self.status_text = widget.status_text

        # Setup table properties
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.preview_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)

        # Connect signals
        widget.browse_template_btn.clicked.connect(self.browse_template)
        widget.browse_source_btn.clicked.connect(self.browse_source)
        widget.fix_export_btn.clicked.connect(self.fix_and_export)
        widget.open_output_btn.clicked.connect(self.open_output_folder)

        # Enable drag and drop on the widget
        widget.setAcceptDrops(True)
        widget.dragEnterEvent = self._drag_enter_event
        widget.dropEvent = self._drop_event

        # Load saved template path
        saved_template = self.app_context.get_setting('report_template_path', '')
        if saved_template and Path(saved_template).exists():
            self.template_path_edit.setText(saved_template)
            self._load_template(saved_template)

        # Populate customer list
        self._populate_customers()

        return widget

    def _get_ui_path(self, relative_path: str) -> Path:
        """Get path to UI file"""
        if getattr(sys, 'frozen', False):
            application_path = Path(sys._MEIPASS)
        else:
            application_path = Path(__file__).parent.parent.parent

        ui_file = application_path / 'modules' / relative_path
        if not ui_file.exists():
            raise FileNotFoundError(f"UI file not found: {ui_file}")
        return ui_file

    def _populate_customers(self):
        """Populate customer dropdown from customer_files_dir"""
        self.customer_combo.clear()
        customers = self.app_context.get_customer_list()
        self.customer_combo.addItems(sorted(customers))

    def _drag_enter_event(self, event):
        """Handle drag enter for file drops"""
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                if url.toLocalFile().lower().endswith(('.xls', '.xlsx')):
                    event.acceptProposedAction()
                    return
        event.ignore()

    def _drop_event(self, event):
        """Handle file drop"""
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.lower().endswith(('.xls', '.xlsx')):
                self._load_source(file_path)
                break

    # ==================== File Loading ====================

    def browse_template(self):
        """Browse for template Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self._widget,
            "Select Template File",
            "",
            "Excel Files (*.xls *.xlsx);;All Files (*.*)"
        )
        if file_path:
            self.template_path_edit.setText(file_path)
            self._load_template(file_path)
            # Save to settings
            self.app_context.set_setting('report_template_path', file_path)
            self.app_context.save_settings()

    def browse_source(self):
        """Browse for source Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self._widget,
            "Select Source Report File",
            "",
            "Excel Files (*.xls *.xlsx);;All Files (*.*)"
        )
        if file_path:
            self._load_source(file_path)

    def _load_template(self, file_path: str):
        """Load template file and extract column names"""
        try:
            df_template = pd.read_excel(file_path, nrows=0)
            self.template_columns = list(df_template.columns)
            self._log(f"Template loaded: {len(self.template_columns)} columns")
            self._update_preview()
        except Exception as e:
            self.show_error("Template Error", f"Failed to load template:\n{str(e)}")
            self.template_columns = None

    def _load_source(self, file_path: str):
        """Load source Excel file"""
        try:
            self.source_df = pd.read_excel(file_path)
            self.source_path_edit.setText(file_path)
            self.source_info_label.setText(
                f"Loaded {len(self.source_df)} rows x {len(self.source_df.columns)} columns"
            )
            self._log(f"Source loaded: {Path(file_path).name}")
            self._update_preview()
        except Exception as e:
            self.show_error("Source Error", f"Failed to load source file:\n{str(e)}")
            self.source_df = None
            self.source_info_label.setText("Failed to load file")

    def _update_preview(self):
        """Update the column mapping preview table"""
        self.preview_table.setRowCount(0)

        if self.template_columns is None or self.source_df is None:
            self.mapping_status_label.setText("Load template and source files to see column mapping")
            self.fix_export_btn.setEnabled(False)
            return

        source_cols = set(self.source_df.columns)
        template_cols = set(self.template_columns)

        matching = source_cols & template_cols
        to_remove = source_cols - template_cols
        to_add = template_cols - source_cols

        # Show all template columns with their status
        for col in self.template_columns:
            row = self.preview_table.rowCount()
            self.preview_table.insertRow(row)

            if col in matching:
                status = "✓"
                action = "Keep (from source)"
                color = Qt.GlobalColor.darkGreen
            else:
                status = "+"
                action = "Add (empty column)"
                color = Qt.GlobalColor.blue

            status_item = QTableWidgetItem(status)
            status_item.setForeground(color)
            self.preview_table.setItem(row, 0, status_item)

            name_item = QTableWidgetItem(col)
            self.preview_table.setItem(row, 1, name_item)

            action_item = QTableWidgetItem(action)
            action_item.setForeground(color)
            self.preview_table.setItem(row, 2, action_item)

        # Show columns to be removed
        for col in sorted(to_remove):
            row = self.preview_table.rowCount()
            self.preview_table.insertRow(row)

            status_item = QTableWidgetItem("✗")
            status_item.setForeground(Qt.GlobalColor.red)
            self.preview_table.setItem(row, 0, status_item)

            name_item = QTableWidgetItem(col)
            self.preview_table.setItem(row, 1, name_item)

            action_item = QTableWidgetItem("Remove (not in template)")
            action_item.setForeground(Qt.GlobalColor.red)
            self.preview_table.setItem(row, 2, action_item)

        self.mapping_status_label.setText(
            f"Matching: {len(matching)} | To Remove: {len(to_remove)} | To Add: {len(to_add)}"
        )

        # Enable export if customer is selected
        self.fix_export_btn.setEnabled(True)

    # ==================== Report Processing ====================

    def fix_and_export(self):
        """Fix the report and export to customer's reports folder"""
        if self.source_df is None or self.template_columns is None:
            self.show_error("Missing Data", "Please load both template and source files first")
            return

        customer = self.customer_combo.currentText().strip()
        if not customer:
            self.show_error("No Customer", "Please select a customer for the output")
            return

        # Get customer files directory
        cf_dir = self.app_context.get_setting('customer_files_dir', '')
        if not cf_dir or not Path(cf_dir).exists():
            self.show_error("Directory Error", "Customer files directory not configured or doesn't exist")
            return

        customer_path = Path(cf_dir) / customer
        if not customer_path.exists():
            self.show_error("Customer Error", f"Customer folder not found:\n{customer_path}")
            return

        # Create reports directory if needed
        reports_dir = customer_path / "reports"
        reports_dir.mkdir(exist_ok=True)

        # Build output filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d")
        output_file = reports_dir / f"{customer}_jobRpt_{timestamp}.xlsx"

        self._log("=" * 50)
        self._log("Starting report transformation...")

        try:
            # Apply transformation
            df_fixed, changed_rows = self._transform_report()

            # Save with formatting
            self._save_formatted_excel(df_fixed, output_file, changed_rows)

            self.last_output_path = output_file
            self.open_output_btn.setEnabled(True)

            self._log("=" * 50)
            self._log(f"SUCCESS: Report saved to:")
            self._log(str(output_file))
            self._log("=" * 50)

            self.show_info("Export Complete",
                          f"Report saved to:\n{output_file}\n\n"
                          f"Schedule changes: {len(changed_rows)}")

        except Exception as e:
            self._log(f"ERROR: {str(e)}")
            self.show_error("Export Failed", f"Failed to export report:\n{str(e)}")

    def _transform_report(self):
        """Transform source data to match template layout"""
        df_fixed = pd.DataFrame()

        # Copy/create columns in template order
        for col in self.template_columns:
            if col in self.source_df.columns:
                df_fixed[col] = self.source_df[col]
            else:
                df_fixed[col] = None
                self._log(f"Added empty column: {col}")

        removed_count = len(set(self.source_df.columns) - set(self.template_columns))
        self._log(f"Removed {removed_count} columns not in template")
        self._log(f"Result: {len(df_fixed)} rows x {len(df_fixed.columns)} columns")

        # Process Scheduled End Date per PO
        changed_rows = []
        if 'Customer PO Number' in df_fixed.columns and 'Scheduled End Date' in df_fixed.columns:
            self._log("Processing Scheduled End Date per PO...")

            # Convert to datetime first (keep as datetime for groupby max to work)
            df_fixed['Scheduled End Date'] = pd.to_datetime(
                df_fixed['Scheduled End Date'], errors='coerce'
            )

            # Set all rows to max date per PO (works with datetime64)
            po_max_dates = df_fixed.groupby('Customer PO Number')['Scheduled End Date'].transform('max')
            df_fixed['Scheduled End Date'] = po_max_dates

            # Now convert to date only (after aggregation)
            df_fixed['Scheduled End Date'] = df_fixed['Scheduled End Date'].dt.date

            unique_pos = df_fixed['Customer PO Number'].nunique()
            self._log(f"Updated dates for {unique_pos} unique POs")

            # Track schedule changes
            changed_rows = self._track_schedule_changes(df_fixed)

        return df_fixed, changed_rows

    def _track_schedule_changes(self, df_fixed):
        """Track schedule changes and add notes"""
        history_file = get_config_dir() / 'schedule_history.json'
        history = {}

        if history_file.exists():
            try:
                with open(history_file, 'r', encoding='utf-8') as f:
                    history = json.load(f)
            except Exception:
                history = {}

        # Ensure Notes column exists
        if 'Notes' not in df_fixed.columns:
            df_fixed['Notes'] = ''

        changed_rows = []
        changes_found = 0

        for idx, row in df_fixed.iterrows():
            po = str(row.get('Customer PO Number', ''))
            line = str(row.get('Line', ''))
            current_date = row.get('Scheduled End Date')

            if not po or pd.isna(current_date):
                continue

            key = f"{po}|{line}"
            current_date_str = str(current_date) if current_date else ''

            if key in history:
                previous_date_str = history[key].get('scheduled_end_date', '')

                if previous_date_str and previous_date_str != current_date_str:
                    changes_found += 1
                    changed_rows.append(idx)

                    existing_notes = str(row.get('Notes', '')) if pd.notna(row.get('Notes')) else ''
                    change_note = f"[{datetime.now().strftime('%m/%d')}] Moved from {previous_date_str}"

                    if existing_notes:
                        new_notes = f"{change_note}; {existing_notes}"
                    else:
                        new_notes = change_note

                    df_fixed.at[idx, 'Notes'] = new_notes

            # Update history
            history[key] = {
                'scheduled_end_date': current_date_str,
                'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'po': po,
                'line': line
            }

        # Save updated history
        try:
            with open(history_file, 'w', encoding='utf-8') as f:
                json.dump(history, f, indent=2)
            self._log(f"History saved ({len(history)} entries)")
        except Exception as e:
            self._log(f"Warning: Could not save history: {e}")

        if changes_found > 0:
            self._log(f"Found {changes_found} schedule changes (will highlight yellow)")

        return changed_rows

    def _save_formatted_excel(self, df_fixed, output_file, changed_rows):
        """Save DataFrame with Excel formatting"""
        # Save initial Excel
        df_fixed.to_excel(output_file, index=False, engine='openpyxl')

        # Load and format
        wb = load_workbook(output_file)
        ws = wb.active

        # Highlight changed Scheduled End Date cells
        if changed_rows:
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # Find Scheduled End Date column
            sched_col = None
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Scheduled End Date':
                    sched_col = col_idx
                    break

            if sched_col:
                for row_idx in changed_rows:
                    excel_row = row_idx + 2  # 1-indexed + header
                    ws.cell(row=excel_row, column=sched_col).fill = yellow_fill

                self._log(f"Highlighted {len(changed_rows)} changed dates")

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add table formatting
        max_row = ws.max_row
        max_col = ws.max_column
        table_ref = f"A1:{ws.cell(max_row, max_col).coordinate}"

        table = Table(displayName="DataTable", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        wb.save(output_file)
        self._log("Applied formatting: auto-fit, frozen header, table style")

    def open_output_folder(self):
        """Open the output folder in file explorer"""
        if self.last_output_path and self.last_output_path.parent.exists():
            success, error = open_folder(str(self.last_output_path.parent))
            if not success:
                self.show_error("Error", error or "Could not open folder")

    def _log(self, message: str):
        """Add message to status log"""
        if self.status_text:
            self.status_text.append(message)
            # Scroll to bottom
            scrollbar = self.status_text.verticalScrollBar()
            scrollbar.setValue(scrollbar.maximum())
        self.log_message(message)

    def cleanup(self):
        """Cleanup resources"""
        pass
